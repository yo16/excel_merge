# openpyxlのテスト
# https://openpyxl.readthedocs.io/en/stable/usage.html

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def main():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'range names'

    for row in range(1,40):
        ws1.append(range(600))
    
    ws2 = wb.create_sheet(title='Pi')
    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title='Data')
    for row in range(10,20):
        for col in range(27,54):
            column_letter = get_column_letter(col)
            _ = ws3.cell(column=col, row=row, \
                value=f'{column_letter}')
    print(ws3['AA10'].value)

    wb.save(filename='./output/test.xlsx')

    return None


if __name__=='__main__':
    main()

